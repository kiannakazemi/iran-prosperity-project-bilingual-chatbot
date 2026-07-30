r"""
Retrieval evaluation harness — augment the golden dataset with retrieval output.

For every golden question this runs one dense candidate pool, reranks it,
expands siblings and section continuations, and reproduces the production
chatbot's answer step. Two retrieval variants are compared side-by-side, and
each variant's final passage set is answered by three LLMs so cost-vs-quality
trade-offs can be measured:

    - Gemini 2.5 Flash       (heavier, more prone to confabulation)
    - Gemini 3.5 Flash-Lite  (production answer model)
    - Qwen-Plus              (DashScope; may refuse political content)

The augmented workbook is written to
``results/retrieval_eval_results.xlsx`` (overwriting any previous run).

For each test, retrieval columns come first, then two columns per model
(answer + end-to-end latency in seconds). Timings include the retrieval work
that variant requires: Test 1 = dense + expansion + answer; Test 2 = dense +
rerank + expansion + answer.

Test 1 — dense-only:
  - "Test1: Dense only"
  - "Test1: Sibling chunks added"
  - "Test1: Final (dense + siblings)"
  - "Test1: Chatbot answer (<model>)"  + "Test1: Timing <model> (seconds)"  ×3

Test 2 — dense + rerank (production):
  - "Test2: Dense + rerank: chunk pool"
  - "Test2: Dense + rerank: chunk pool reranked"
  - "Test2: Dense + rerank: top 5 chunks selected"
  - "Test2: Sibling chunks added"
  - "Test2: Final (rerank + siblings + continuations)"
  - "Test2: Chatbot answer (<model>)"  + "Test2: Timing <model> (seconds)"  ×3

Both variants run split-sibling expansion and section-continuation stitching,
matching ``ReliableRetriever.query``. Reranking is section-aware: candidates are
sent to Cohere with their header trail prepended, so the reranker can tell apart
same-named sections ("Key Priorities") repeated under different phases.

A third variant — LLM header-sibling routing — was evaluated and removed. It
changed the retrieved set on only 5/47 English and 9/47 Persian questions, so
most apparent differences were model nondeterminism rather than retrieval; and
every chunk it usefully added was already inside the dense pool, merely ranked
below top-5, which section-aware reranking handles without an extra LLM call.
See ``EVALUATION.md``.

Notes:
  - Qwen-Plus may refuse politically sensitive content with an
    "inappropriate content" error — refusals are surfaced as
    "[Qwen refused: ...]" in the answer column so the pattern is visible.
  - Both Gemini variants use the same GEMINI_API_KEY (per-model quotas).

Run (with the API stopped, so Qdrant is free):
    cd rag_pipeline
    ..\.venv\Scripts\python.exe -m indexing.retrieval.eval.run_eval
    ..\.venv\Scripts\python.exe -m indexing.retrieval.eval.run_eval --limit 3
"""

from __future__ import annotations

import argparse
import time
from pathlib import Path

import pandas as pd

from indexing.retrieval import RERANK_CANDIDATE_POOL, ReliableRetriever, RetrievalResult
from chatbot.engine import (
    GEMINI_MODEL,
    MAX_OUTPUT_TOKENS,
    _gemini_generate,
    _qwen_complete,
    _system_prompt,
)

# Model identifiers used by generate_answer(). GEMINI_MODEL (from engine.py)
# is the production default; Flash-Lite is its cheaper sibling.
# Using the newest stable Flash-Lite (July 2026 release) available on the
# project's API key — has `thinking: true` and larger output-token limit.
GEMINI_FLASH_MODEL = GEMINI_MODEL              # "gemini-2.5-flash"
GEMINI_FLASH_LITE_MODEL = "gemini-3.5-flash-lite"

MODEL_KINDS = ("gemini_flash", "gemini_flash_lite", "qwen")
ALL_MODEL_KINDS = MODEL_KINDS

# Which tests to generate answers for. Narrowed by --tests so a single variant
# can be re-measured without paying for the others.
ALL_TESTS = ("1", "2")
TESTS = ALL_TESTS
MODEL_LABEL = {
    "gemini_flash": "Gemini Flash",
    "gemini_flash_lite": "Gemini Flash-Lite",
    "qwen": "Qwen",
}

HERE = Path(__file__).resolve().parent
DATA_DIR = HERE / "data"
RESULTS_DIR = HERE / "results"

# Final passage count both variants are cut to (matches engine.TOP_K).
FINAL_TOP_K = 5

# Retrieval columns per test (unchanged from previous iteration).
# Test 1 (dense-only)
COL_DENSE = "Test1: Dense only"
COL_DENSE_SIBLINGS = "Test1: Sibling chunks added"
COL_DENSE_FINAL = "Test1: Final (dense + siblings)"

# Test 2 (dense + rerank)
COL_POOL = "Test2: Dense + rerank: chunk pool"
COL_POOL_RERANKED = "Test2: Dense + rerank: chunk pool reranked"
COL_TOP5 = "Test2: Dense + rerank: top 5 chunks selected"
COL_RERANK_SIBLINGS = "Test2: Sibling chunks added"
COL_RERANK_FINAL = "Test2: Final (rerank + siblings + continuations)"




def _answer_col(test: str, model: str) -> str:
    return f"Test{test}: Chatbot answer ({MODEL_LABEL[model]})"


def _timing_col(test: str, model: str) -> str:
    return f"Test{test}: Timing {MODEL_LABEL[model]} (seconds)"


# Answer/timing columns per test × model. Rebuilt by _build_new_cols() after the
# CLI narrows MODEL_KINDS / TESTS, so selecting a subset does not leave stale
# column names behind.
COL_T1_ANSWER: dict[str, str] = {}
COL_T1_TIMING: dict[str, str] = {}
COL_T2_ANSWER: dict[str, str] = {}
COL_T2_TIMING: dict[str, str] = {}


def _test_block_cols(retrieval_cols: list[str],
                     answer_cols: dict, timing_cols: dict) -> list[str]:
    """For one test: retrieval cols first, then answer+timing per model."""
    out = list(retrieval_cols)
    for m in MODEL_KINDS:
        out.append(answer_cols[m])
        out.append(timing_cols[m])
    return out


def _build_new_cols() -> list[str]:
    """Compute the columns to append, honouring the current TESTS/MODEL_KINDS.

    Called once from main() after argument parsing. Mutates the module-level
    per-test answer/timing dicts so the run body can index them directly.
    """
    global COL_T1_ANSWER, COL_T1_TIMING, COL_T2_ANSWER, COL_T2_TIMING

    COL_T1_ANSWER = {m: _answer_col("1", m) for m in MODEL_KINDS}
    COL_T1_TIMING = {m: _timing_col("1", m) for m in MODEL_KINDS}
    COL_T2_ANSWER = {m: _answer_col("2", m) for m in MODEL_KINDS}
    COL_T2_TIMING = {m: _timing_col("2", m) for m in MODEL_KINDS}

    cols: list[str] = []
    if "1" in TESTS:
        cols += _test_block_cols(
            [COL_DENSE, COL_DENSE_SIBLINGS, COL_DENSE_FINAL],
            COL_T1_ANSWER, COL_T1_TIMING,
        )
    if "2" in TESTS:
        cols += _test_block_cols(
            [COL_POOL, COL_POOL_RERANKED, COL_TOP5,
             COL_RERANK_SIBLINGS, COL_RERANK_FINAL],
            COL_T2_ANSWER, COL_T2_TIMING,
        )
    return cols


NEW_COLS: list[str] = []


# ── dataset column resolution ──────────────────────────────────────────
QUESTION_COLS = ["Question", "question", "سؤال", "سوال", "پرسش"]


def _question_col(df: pd.DataFrame):
    for c in QUESTION_COLS:
        if c in df.columns:
            return c
    return None


def _lang_for_sheet(sheet_name: str) -> str:
    return "en" if "english" in sheet_name.lower() else "fa"


# ── formatting helpers ─────────────────────────────────────────────────
def clabel(cid: str) -> str:
    """'en__…__00005' → 'chunk_5'."""
    if not cid:
        return "chunk_?"
    n = cid.split("__")[-1]
    try:
        n = str(int(n))
    except ValueError:
        pass
    return f"chunk_{n}"


def fmt_scored(hits: list) -> str:
    """One per line: 'chunk_5: 0.612'."""
    return "\n".join(f"{clabel(h.chunk_id)}: {h.score:.3f}" for h in hits)


def fmt_ids(hits: list) -> str:
    """Comma-separated chunk ids: 'chunk_5, chunk_6, chunk_7'."""
    return ", ".join(clabel(h.chunk_id) for h in hits)


def _cid_of(x) -> str:
    """Extract chunk_id from either a _Hit or a RetrievalResult."""
    if hasattr(x, "chunk_id"):
        return x.chunk_id
    return x.metadata.get("chunk_id", "")


def _fmt_siblings(sibs) -> str:
    """Comma-separated sibling chunk ids. Empty string when no siblings were
    added (i.e. no split-part chunks in the top-K)."""
    return ", ".join(clabel(_cid_of(s)) for s in sibs)


def _fmt_final(expanded) -> str:
    """Final passage list = top-K primaries followed by any additions.
    Split-part siblings are marked '(sibling)'; LLM-router siblings are
    marked '(header)' so the two kinds are scannable in Excel."""
    parts: list[str] = []
    for x in expanded:
        cid = clabel(_cid_of(x))
        src = x.source
        if "header_sibling" in src:
            parts.append(f"{cid} (header)")
        elif "sibling" in src:
            parts.append(f"{cid} (sibling)")
        else:
            parts.append(cid)
    return ", ".join(parts)


def _build_context(question: str, results: list) -> str:
    """Assemble the "CONTEXT:\n...\n\nQUESTION:\n..." user message.

    Must stay byte-identical to ``ChatEngine._build_context`` plus the
    CONTEXT/QUESTION wrapper, or the eval measures a pipeline the app does not
    run. It previously omitted the ``[Section: ...]`` heading trail that
    production sends, which understated results on phase-scoped questions —
    exactly the questions the heading trail exists to fix.
    """
    parts = []
    for i, r in enumerate(results, 1):
        wp = r.metadata.get("white_paper", "Unknown")
        try:
            ps = int(r.metadata.get("page_start", 0))
            pe = int(r.metadata.get("page_end", 0))
        except (TypeError, ValueError):
            ps = pe = 0
        page_info = f"p.{ps}" if ps == pe else f"pp.{ps}-{pe}"

        header = (r.metadata.get("header_path") or "").strip("/")
        section = f"\n[Section: {header.replace('/', ' > ')}]" if header else ""

        parts.append(
            f"[Passage {i} — Source: {wp}, {page_info}]{section}\n{r.text}"
        )
    context = "\n\n---\n\n".join(parts)
    return f"CONTEXT:\n{context}\n\nQUESTION:\n{question}"


def generate_answer(
    question: str,
    results: list,
    lang: str,
    model_kind: str,
    *,
    system_prompt: str | None = None,
) -> str:
    """Reproduce the chatbot's answer step for a set of retrieved passages,
    routed to a specific model (``gemini_flash``, ``gemini_flash_lite``, or
    ``qwen``). Same context format and max_tokens as ``chatbot/engine.py``.

    Pass ``system_prompt`` to override the default if needed.

    Failures return a short "[<label> error: ...]" string so a single model's
    outage doesn't kill the whole eval row.
    """
    if not results:
        return ""
    user = _build_context(question, results)
    system = system_prompt if system_prompt is not None else _system_prompt(lang)
    max_tok = MAX_OUTPUT_TOKENS.get(lang, 1536)
    try:
        if model_kind == "gemini_flash":
            answer = _gemini_generate(system, user, max_tokens=max_tok,
                                      model=GEMINI_FLASH_MODEL)
        elif model_kind == "gemini_flash_lite":
            answer = _gemini_generate(system, user, max_tokens=max_tok,
                                      model=GEMINI_FLASH_LITE_MODEL)
        elif model_kind == "qwen":
            # Direct Qwen call (no Gemini fallback) so refusals are visible.
            answer = _qwen_complete(system, user, max_tokens=max_tok)
        else:
            return f"[unknown model kind: {model_kind}]"
        return answer or "(No response generated)"
    except Exception as e:
        # Qwen moderation refusals surface as RuntimeError with 400 + message.
        msg = str(e)
        if model_kind == "qwen" and (
            "inappropriate content" in msg.lower()
            or "data_inspection_failed" in msg.lower()
            or "400" in msg
        ):
            return f"[Qwen refused: {msg[:140]}]"
        return f"[{MODEL_LABEL[model_kind]} error: {msg[:140]}]"


# ── Excel styling (wrap the multi-line / long columns) ─────────────────
_TIMING_COLS = (
    set(COL_T1_TIMING.values())
    | set(COL_T2_TIMING.values())
)
_WRAP_COLS = set(NEW_COLS) - _TIMING_COLS


def _style(writer) -> None:
    from openpyxl.styles import Alignment, Font

    top_wrap = Alignment(wrap_text=True, vertical="top")
    for ws in writer.book.worksheets:
        ws.freeze_panes = "A2"
        for c in ws[1]:
            c.font = Font(bold=True)
        headers = {c.value: c.column_letter for c in ws[1]}
        for name, col in headers.items():
            if name in _TIMING_COLS:
                ws.column_dimensions[col].width = 10
            elif name in _WRAP_COLS:
                ws.column_dimensions[col].width = 40
            else:
                ws.column_dimensions[col].width = 22
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = top_wrap


# ── dataset loading ────────────────────────────────────────────────────
def find_data_file(explicit: str | None) -> Path:
    if explicit:
        return Path(explicit)
    xls = sorted(DATA_DIR.glob("*.xlsx"))
    if not xls:
        raise SystemExit(f"No .xlsx found in {DATA_DIR}. Put the golden dataset there.")
    return xls[0]


# ── main ───────────────────────────────────────────────────────────────
def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--file", default=None, help="path to golden dataset xlsx")
    ap.add_argument("--limit", type=int, default=0, help="max questions per sheet (0=all)")
    ap.add_argument("--sheet", default=None, help="only run this sheet name")
    ap.add_argument(
        "--tests", default=",".join(ALL_TESTS),
        help='comma-separated tests to run, e.g. "4" or "2,4" (default: all)',
    )
    ap.add_argument(
        "--models", default=",".join(ALL_MODEL_KINDS),
        help=('comma-separated models, e.g. "gemini_flash_lite" '
              f'(choices: {",".join(ALL_MODEL_KINDS)})'),
    )
    args = ap.parse_args()

    global TESTS, MODEL_KINDS, NEW_COLS
    TESTS = tuple(t.strip() for t in args.tests.split(",") if t.strip() in ALL_TESTS)
    if not TESTS:
        ap.error(f"--tests must name at least one of {ALL_TESTS}")
    MODEL_KINDS = tuple(
        m.strip() for m in args.models.split(",") if m.strip() in ALL_MODEL_KINDS
    )
    if not MODEL_KINDS:
        ap.error(f"--models must name at least one of {ALL_MODEL_KINDS}")
    NEW_COLS = _build_new_cols()
    print(f"Tests:  {', '.join(TESTS)}")
    print(f"Models: {', '.join(MODEL_KINDS)}")

    data_file = find_data_file(args.file)
    print(f"Dataset: {data_file}")
    book = pd.read_excel(data_file, sheet_name=None)

    r = ReliableRetriever()

    RESULTS_DIR.mkdir(parents=True, exist_ok=True)

    out_sheets: dict[str, pd.DataFrame] = {}

    for sheet_name, df in book.items():
        if args.sheet and sheet_name != args.sheet:
            out_sheets[sheet_name[:31]] = df
            continue
        qcol = _question_col(df)
        if qcol is None:
            print(f"  ! sheet {sheet_name!r}: no Question column, leaving as-is")
            out_sheets[sheet_name[:31]] = df
            continue

        lang = _lang_for_sheet(sheet_name)
        if args.limit:
            df = df.head(args.limit)
        print(f"\n=== {sheet_name} (lang={lang}, {len(df)} questions) ===")

        cols = {c: [] for c in NEW_COLS}
        for _, row in df.iterrows():
            q = str(row[qcol]).strip()
            if not q or q.lower() == "nan":
                for c in NEW_COLS:
                    cols[c].append("")
                continue

            _progress_col = NEW_COLS[0]
            print(f"  [{len(cols[_progress_col]) + 1}] {q[:65]}")

            # ── Shared retrieval steps: dense pool + rerank
            t0 = time.time()
            pool = r.dense_search(q, lang, RERANK_CANDIDATE_POOL)
            t_dense = time.time() - t0

            # rerank_pool sends section-aware documents (header trail + topic),
            # so the reranker can distinguish same-named sections across phases.
            # Skipped entirely when Test 2 is not selected — it is a paid call.
            reranked: list = []
            t_rerank = 0.0
            if "2" in TESTS:
                t0 = time.time()
                reranked = r.rerank_pool(q, pool, len(pool))
                t_rerank = time.time() - t0

            # ── Test 1 retrieval: top 5 dense → sibling + continuation expansion
            t0 = time.time()
            t1_top = pool[:FINAL_TOP_K]
            t1_expanded = r._expand_split_siblings([
                RetrievalResult(text=h.text, metadata=h.metadata,
                                rerank_score=h.score, source="dense")
                for h in t1_top
            ])
            t1_expanded = r._expand_continuations(t1_expanded)
            t_t1_expand = time.time() - t0
            t1_siblings = [x for x in t1_expanded
                           if "sibling" in x.source or "continuation" in x.source]

            # ── Test 2 retrieval: top 5 reranked → sibling + continuation expansion
            t0 = time.time()
            t2_top = reranked[:FINAL_TOP_K]
            t2_expanded = r._expand_split_siblings([
                RetrievalResult(text=h.text, metadata=h.metadata,
                                rerank_score=h.score, source="rerank")
                for h in t2_top
            ])
            t2_expanded = r._expand_continuations(t2_expanded)
            t_t2_expand = time.time() - t0
            t2_siblings = [x for x in t2_expanded
                           if "sibling" in x.source or "continuation" in x.source]

            # Each test's retrieval-prefix latency (added to each model's answer time).
            t1_retrieval = t_dense + t_t1_expand
            t2_retrieval = t_dense + t_rerank + t_t2_expand

            t1_answers: dict[str, tuple[str, float]] = {}
            t2_answers: dict[str, tuple[str, float]] = {}
            for m in MODEL_KINDS:
                if "1" in TESTS:
                    t0 = time.time()
                    a1 = generate_answer(q, t1_expanded, lang, m)
                    t1_answers[m] = (a1, t1_retrieval + (time.time() - t0))

                if "2" in TESTS:
                    t0 = time.time()
                    a2 = generate_answer(q, t2_expanded, lang, m)
                    t2_answers[m] = (a2, t2_retrieval + (time.time() - t0))


            # ── Test 1 columns
            if "1" in TESTS:
                cols[COL_DENSE].append(fmt_scored(t1_top))
                cols[COL_DENSE_SIBLINGS].append(_fmt_siblings(t1_siblings))
                cols[COL_DENSE_FINAL].append(_fmt_final(t1_expanded))
                for m in MODEL_KINDS:
                    ans, sec = t1_answers[m]
                    cols[COL_T1_ANSWER[m]].append(ans)
                    cols[COL_T1_TIMING[m]].append(f"{sec:.2f}")

            # ── Test 2 columns
            if "2" in TESTS:
                cols[COL_POOL].append(fmt_ids(pool))
                cols[COL_POOL_RERANKED].append(fmt_ids(reranked))
                cols[COL_TOP5].append(fmt_scored(t2_top))
                cols[COL_RERANK_SIBLINGS].append(_fmt_siblings(t2_siblings))
                cols[COL_RERANK_FINAL].append(_fmt_final(t2_expanded))
                for m in MODEL_KINDS:
                    ans, sec = t2_answers[m]
                    cols[COL_T2_ANSWER[m]].append(ans)
                    cols[COL_T2_TIMING[m]].append(f"{sec:.2f}")

        df = df.copy()
        for c in NEW_COLS:
            df[c] = cols[c]
        out_sheets[sheet_name[:31]] = df

    out_path = RESULTS_DIR / "retrieval_eval_results.xlsx"
    with pd.ExcelWriter(out_path) as w:
        for name, df in out_sheets.items():
            df.to_excel(w, sheet_name=name, index=False)
        _style(w)

    print(f"\nWrote:\n  {out_path}")


if __name__ == "__main__":
    main()
